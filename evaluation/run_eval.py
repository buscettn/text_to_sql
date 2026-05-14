import os
import asyncio
import pandas as pd
import uuid
import warnings
from datetime import datetime
from sdk import TextToSQLEngine, SQLRequest, SQLResponse
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
import nltk

# Suppress openpyxl warnings about data validation
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Ensure nltk resources are available for tokenization
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt', quiet=True)

async def run_query(engine: TextToSQLEngine, query: str, thread_id: str) -> SQLResponse:
    """Runs a single query through the TextToSQLEngine."""
    request = SQLRequest(query=query, thread_id=thread_id)
    try:
        response = await engine.ainvoke(request)
        return response
    except Exception as e:
        return SQLResponse(
            sql=None,
            message=str(e),
            status="error",
            thread_id=thread_id
        )

async def main():
    # Paths
    ground_truth_path = "evaluation/ground_truth.xlsx"
    results_dir = "evaluation/results"
    os.makedirs(results_dir, exist_ok=True)
    
    print(f"--- Text-to-SQL Evaluation ---")
    
    # Load ground truth
    if not os.path.exists(ground_truth_path):
        print(f"Error: {ground_truth_path} not found.")
        return

    try:
        df_gt = pd.read_excel(ground_truth_path, sheet_name=0)
    except Exception as e:
        print(f"Error reading {ground_truth_path}: {e}")
        return

    # Verify columns
    if 'query' not in df_gt.columns or 'sql' not in df_gt.columns:
        print("Error: Missing 'query' or 'sql' columns in the first sheet.")
        return

    # Filter invalid rows (missing query or sql)
    valid_mask = df_gt['query'].notna() & df_gt['sql'].notna()
    skipped_rows_count = len(df_gt) - valid_mask.sum()
    if skipped_rows_count > 0:
        print(f"Warning: Skipping {skipped_rows_count} rows with missing query or sql.")
    
    df_eval = df_gt[valid_mask].copy()
    
    if df_eval.empty:
        print("Error: No valid rows found for evaluation.")
        return

    engine = TextToSQLEngine()
    semaphore = asyncio.Semaphore(2)  # Max concurrency as per plan
    
    async def task(row_idx, row_data):
        async with semaphore:
            thread_id = str(uuid.uuid4())
            response = await run_query(engine, row_data['query'], thread_id)
            return row_idx, response

    print(f"Processing {len(df_eval)} queries with max concurrency 2...")
    execution_tasks = [task(i, row) for i, row in df_eval.iterrows()]
    results = await asyncio.gather(*execution_tasks)
    
    # Collect results and calculate metrics
    detail_data = []
    smoothing = SmoothingFunction().method1
    
    for idx, response in results:
        gt_query = df_eval.loc[idx, 'query']
        gt_sql = df_eval.loc[idx, 'sql']
        
        gen_sql = response.sql if response.status == "success" else "FAILED"
        status = response.status
        mapped_domain = response.data_domain or "unknown"
        error_msg = response.message if response.status != "success" else ""
        
        # Literal comparison (strict as per plan)
        is_equal = (gen_sql == gt_sql)
        
        # BLEU Score (Simple whitespace tokenization)
        # Handle cases where SQL is missing or empty
        ref_tokens = str(gt_sql).split()
        cand_tokens = str(gen_sql).split() if response.status == "success" else []
        
        if cand_tokens:
            bleu = sentence_bleu([ref_tokens], cand_tokens, smoothing_function=smoothing)
        else:
            bleu = 0.0
            
        detail_data.append({
            'query': gt_query,
            'ground_truth_sql': gt_sql,
            'result': gen_sql,
            'mapped domain': mapped_domain,
            'equal flag': is_equal,
            'bleu score': bleu,
            'error message': error_msg,
            'status': status
        })

    df_details = pd.DataFrame(detail_data)
    
    # Aggregate Metrics for Overview
    total_valid = len(df_details)
    successful_runs = df_details[df_details['status'] == "success"]
    failed_runs_count = total_valid - len(successful_runs)
    
    accuracy = df_details['equal flag'].mean()
    bleu_scores = df_details['bleu score']
    
    overview_data = {
        'Metric': [
            'Accuracy (percentage/mean)',
            'BLEU score (mean)',
            'BLEU score (median)',
            'BLEU score (p80)',
            'BLEU score (p90)',
            'Status != success (%)',
            'Status != success (count)',
            'Number of skipped/invalid rows'
        ],
        'Value': [
            accuracy,
            bleu_scores.mean(),
            bleu_scores.median(),
            bleu_scores.quantile(0.8),
            bleu_scores.quantile(0.9),
            (failed_runs_count / total_valid) if total_valid > 0 else 0,
            failed_runs_count,
            skipped_rows_count
        ]
    }
    df_overview = pd.DataFrame(overview_data)
    
    # Output File Generation
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M")
    output_filename = f"evaluation_{timestamp}.xlsx"
    output_path = os.path.join(results_dir, output_filename)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_overview.to_excel(writer, sheet_name='Overview', index=False)
            # Drop the helper status column for the final report
            df_details.drop(columns=['status']).to_excel(writer, sheet_name='Detail', index=False)
            
            # Formatting: Green for equal flag = True
            from openpyxl.styles import PatternFill
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Light green
            
            workbook = writer.book
            if 'Detail' in workbook.sheetnames:
                detail_sheet = workbook['Detail']
                
                # Find column index for 'equal flag'
                equal_col_idx = None
                for cell in detail_sheet[1]:
                    if cell.value == 'equal flag':
                        equal_col_idx = cell.column
                        break
                
                if equal_col_idx:
                    for row in range(2, len(df_details) + 2):
                        cell = detail_sheet.cell(row=row, column=equal_col_idx)
                        if cell.value is True:
                            cell.fill = green_fill
        
        print(f"Evaluation complete. Results saved to: {output_path}")
        
    except Exception as e:
        print(f"Error saving results to Excel: {e}")

if __name__ == "__main__":
    asyncio.run(main())
